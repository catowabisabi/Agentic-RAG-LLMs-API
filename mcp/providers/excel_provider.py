"""
Excel Provider

基於 excel-mcp-server 的 Excel 文件操作封裝。
提供 LLM 可用的 Excel 讀寫、格式化、公式等功能。

依賴: openpyxl>=3.1.5
"""

import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from mcp.providers.base_provider import BaseProvider

logger = logging.getLogger(__name__)


class ExcelProviderError(Exception):
    """Excel Provider 相關錯誤"""
    pass


class ExcelProvider(BaseProvider):
    """
    Excel 文件操作 Provider
    
    功能：
    - 創建/讀取 Excel 文件
    - 讀寫數據
    - 應用公式
    - 格式化（字體、顏色、邊框）
    - 工作表管理
    """
    
    def __init__(self, base_path: Optional[str] = None):
        """
        初始化 Excel Provider
        
        Args:
            base_path: Excel 文件的基礎路徑（可選）
                      如果提供，所有相對路徑會基於此路徑解析
        """
        super().__init__(
            name="excel_provider",
            description="Excel file operations: read, write, format, formulas"
        )
        
        # 設定基礎路徑
        self.base_path = Path(base_path) if base_path else Path.cwd()
        if not self.base_path.exists():
            self.base_path.mkdir(parents=True, exist_ok=True)
            
        logger.info(f"ExcelProvider initialized with base_path: {self.base_path}")
    
    def _resolve_path(self, filepath: str) -> Path:
        """解析文件路徑（支持相對和絕對路徑）"""
        path = Path(filepath)
        if path.is_absolute():
            return path
        return self.base_path / path
    
    # ========================================
    # 工作簿操作
    # ========================================
    
    def create_workbook(
        self, 
        filepath: str, 
        sheet_name: str = "Sheet1"
    ) -> Dict[str, Any]:
        """
        創建新的 Excel 工作簿
        
        Args:
            filepath: 文件路徑
            sheet_name: 初始工作表名稱
            
        Returns:
            {"message": str, "filepath": str, "sheet_name": str}
        """
        try:
            full_path = self._resolve_path(filepath)
            full_path.parent.mkdir(parents=True, exist_ok=True)
            
            wb = Workbook()
            # 修改默認工作表名稱
            ws = wb.active
            ws.title = sheet_name
            
            wb.save(str(full_path))
            wb.close()
            
            logger.info(f"Created workbook: {full_path}")
            return {
                "message": f"Created workbook successfully",
                "filepath": str(full_path),
                "sheet_name": sheet_name
            }
        except Exception as e:
            logger.error(f"Failed to create workbook: {e}")
            raise ExcelProviderError(f"Failed to create workbook: {e}")
    
    def get_workbook_info(self, filepath: str) -> Dict[str, Any]:
        """
        獲取工作簿資訊
        
        Args:
            filepath: 文件路徑
            
        Returns:
            {"filename": str, "sheets": List[str], "size": int, ...}
        """
        try:
            full_path = self._resolve_path(filepath)
            if not full_path.exists():
                raise ExcelProviderError(f"File not found: {filepath}")
            
            wb = load_workbook(str(full_path), read_only=True)
            
            info = {
                "filename": full_path.name,
                "filepath": str(full_path),
                "sheets": wb.sheetnames,
                "size": full_path.stat().st_size,
                "modified": full_path.stat().st_mtime
            }
            
            wb.close()
            return info
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to get workbook info: {e}")
            raise ExcelProviderError(f"Failed to get workbook info: {e}")
    
    # ========================================
    # 工作表操作
    # ========================================
    
    def create_sheet(
        self, 
        filepath: str, 
        sheet_name: str
    ) -> Dict[str, str]:
        """
        創建新工作表
        
        Args:
            filepath: 文件路徑
            sheet_name: 工作表名稱
            
        Returns:
            {"message": str}
        """
        try:
            full_path = self._resolve_path(filepath)
            wb = load_workbook(str(full_path))
            
            if sheet_name in wb.sheetnames:
                wb.close()
                raise ExcelProviderError(f"Sheet '{sheet_name}' already exists")
            
            wb.create_sheet(sheet_name)
            wb.save(str(full_path))
            wb.close()
            
            return {"message": f"Sheet '{sheet_name}' created successfully"}
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to create sheet: {e}")
            raise ExcelProviderError(f"Failed to create sheet: {e}")
    
    def delete_sheet(
        self, 
        filepath: str, 
        sheet_name: str
    ) -> Dict[str, str]:
        """刪除工作表"""
        try:
            full_path = self._resolve_path(filepath)
            wb = load_workbook(str(full_path))
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ExcelProviderError(f"Sheet '{sheet_name}' not found")
            
            del wb[sheet_name]
            wb.save(str(full_path))
            wb.close()
            
            return {"message": f"Sheet '{sheet_name}' deleted successfully"}
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to delete sheet: {e}")
            raise ExcelProviderError(f"Failed to delete sheet: {e}")
    
    # ========================================
    # 數據讀寫
    # ========================================
    
    def read_range(
        self,
        filepath: str,
        sheet_name: str,
        start_cell: str = "A1",
        end_cell: Optional[str] = None
    ) -> List[List[Any]]:
        """
        讀取單元格範圍數據
        
        Args:
            filepath: 文件路徑
            sheet_name: 工作表名稱
            start_cell: 起始單元格 (如 "A1")
            end_cell: 結束單元格 (如 "D10")，如果為 None 則讀取整個工作表
            
        Returns:
            二維數組，每行是一個列表
        """
        try:
            full_path = self._resolve_path(filepath)
            wb = load_workbook(str(full_path), read_only=True)
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ExcelProviderError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # 解析範圍
            if end_cell:
                range_str = f"{start_cell}:{end_cell}"
            else:
                # 讀取整個已使用範圍
                max_col = get_column_letter(ws.max_column)
                range_str = f"A1:{max_col}{ws.max_row}"
            
            # 讀取數據
            data = []
            for row in ws[range_str]:
                row_data = [cell.value for cell in row]
                data.append(row_data)
            
            wb.close()
            return data
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to read range: {e}")
            raise ExcelProviderError(f"Failed to read range: {e}")
    
    def write_data(
        self,
        filepath: str,
        sheet_name: str,
        data: List[List[Any]],
        start_cell: str = "A1"
    ) -> Dict[str, str]:
        """
        寫入數據到工作表
        
        Args:
            filepath: 文件路徑
            sheet_name: 工作表名稱
            data: 二維數組數據
            start_cell: 起始單元格
            
        Returns:
            {"message": str}
        """
        try:
            full_path = self._resolve_path(filepath)
            
            # 如果文件不存在，創建新文件
            if not full_path.exists():
                self.create_workbook(str(full_path), sheet_name)
                
            wb = load_workbook(str(full_path))
            
            # 如果工作表不存在，創建
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            
            ws = wb[sheet_name]
            
            # 解析起始單元格
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            col_letter, row = coordinate_from_string(start_cell)
            start_row = row
            start_col = column_index_from_string(col_letter)
            
            # 寫入數據
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    ws.cell(
                        row=start_row + row_idx,
                        column=start_col + col_idx,
                        value=value
                    )
            
            wb.save(str(full_path))
            wb.close()
            
            return {"message": f"Data written to '{sheet_name}' successfully"}
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to write data: {e}")
            raise ExcelProviderError(f"Failed to write data: {e}")
    
    # ========================================
    # 公式操作
    # ========================================
    
    def apply_formula(
        self,
        filepath: str,
        sheet_name: str,
        cell: str,
        formula: str
    ) -> Dict[str, Any]:
        """
        應用 Excel 公式到單元格
        
        Args:
            filepath: 文件路徑
            sheet_name: 工作表名稱
            cell: 單元格地址 (如 "A1")
            formula: Excel 公式 (如 "=SUM(A1:A10)")
            
        Returns:
            {"message": str, "cell": str, "formula": str}
        """
        try:
            full_path = self._resolve_path(filepath)
            wb = load_workbook(str(full_path))
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ExcelProviderError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # 確保公式以 = 開頭
            if not formula.startswith("="):
                formula = "=" + formula
            
            ws[cell] = formula
            wb.save(str(full_path))
            wb.close()
            
            return {
                "message": f"Formula applied to {cell}",
                "cell": cell,
                "formula": formula
            }
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to apply formula: {e}")
            raise ExcelProviderError(f"Failed to apply formula: {e}")
    
    # ========================================
    # 格式化操作
    # ========================================
    
    def format_cells(
        self,
        filepath: str,
        sheet_name: str,
        cell_range: str,
        font_name: Optional[str] = None,
        font_size: Optional[int] = None,
        font_bold: Optional[bool] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
        border: Optional[bool] = None
    ) -> Dict[str, str]:
        """
        格式化單元格
        
        Args:
            filepath: 文件路徑
            sheet_name: 工作表名稱
            cell_range: 單元格範圍 (如 "A1:D10")
            font_name: 字體名稱 (如 "Arial")
            font_size: 字體大小
            font_bold: 是否粗體
            font_color: 字體顏色 (HEX, 如 "FF0000")
            bg_color: 背景顏色 (HEX, 如 "FFFF00")
            border: 是否添加邊框
            
        Returns:
            {"message": str}
        """
        try:
            full_path = self._resolve_path(filepath)
            wb = load_workbook(str(full_path))
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise ExcelProviderError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # 創建樣式
            font = None
            if any([font_name, font_size, font_bold, font_color]):
                font = Font(
                    name=font_name or "Calibri",
                    size=font_size or 11,
                    bold=font_bold or False,
                    color=font_color
                )
            
            fill = None
            if bg_color:
                fill = PatternFill(
                    start_color=bg_color,
                    end_color=bg_color,
                    fill_type="solid"
                )
            
            border_style = None
            if border:
                side = Side(style="thin", color="000000")
                border_style = Border(
                    left=side, right=side, top=side, bottom=side
                )
            
            # 應用格式
            for row in ws[cell_range]:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if border_style:
                        cell.border = border_style
            
            wb.save(str(full_path))
            wb.close()
            
            return {"message": f"Formatting applied to {cell_range}"}
            
        except ExcelProviderError:
            raise
        except Exception as e:
            logger.error(f"Failed to format cells: {e}")
            raise ExcelProviderError(f"Failed to format cells: {e}")
    
    # ========================================
    # MCP Provider 接口實現
    # ========================================
    
    async def execute(
        self, 
        operation: str, 
        **kwargs
    ) -> Dict[str, Any]:
        """
        執行 Excel 操作
        
        Args:
            operation: 操作名稱
            **kwargs: 操作參數
            
        Returns:
            操作結果
        """
        operations = {
            "create_workbook": self.create_workbook,
            "get_workbook_info": self.get_workbook_info,
            "create_sheet": self.create_sheet,
            "delete_sheet": self.delete_sheet,
            "read_range": self.read_range,
            "write_data": self.write_data,
            "apply_formula": self.apply_formula,
            "format_cells": self.format_cells,
        }
        
        if operation not in operations:
            raise ExcelProviderError(f"Unknown operation: {operation}")
        
        try:
            result = operations[operation](**kwargs)
            return {
                "success": True,
                "operation": operation,
                "result": result
            }
        except Exception as e:
            logger.error(f"Excel operation '{operation}' failed: {e}")
            return {
                "success": False,
                "operation": operation,
                "error": str(e)
            }
    
    def get_available_operations(self) -> List[Dict[str, Any]]:
        """返回可用操作列表"""
        return [
            {
                "name": "create_workbook",
                "description": "Create a new Excel workbook",
                "parameters": ["filepath", "sheet_name"]
            },
            {
                "name": "get_workbook_info",
                "description": "Get workbook information",
                "parameters": ["filepath"]
            },
            {
                "name": "create_sheet",
                "description": "Create a new worksheet",
                "parameters": ["filepath", "sheet_name"]
            },
            {
                "name": "delete_sheet",
                "description": "Delete a worksheet",
                "parameters": ["filepath", "sheet_name"]
            },
            {
                "name": "read_range",
                "description": "Read data from cell range",
                "parameters": ["filepath", "sheet_name", "start_cell", "end_cell"]
            },
            {
                "name": "write_data",
                "description": "Write data to worksheet",
                "parameters": ["filepath", "sheet_name", "data", "start_cell"]
            },
            {
                "name": "apply_formula",
                "description": "Apply Excel formula to cell",
                "parameters": ["filepath", "sheet_name", "cell", "formula"]
            },
            {
                "name": "format_cells",
                "description": "Format cells (font, color, border)",
                "parameters": ["filepath", "sheet_name", "cell_range", "font_name", "font_size", "font_bold", "font_color", "bg_color", "border"]
            }
        ]
