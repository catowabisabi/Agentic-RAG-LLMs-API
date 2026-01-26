"""
File Control MCP Provider

Handles Excel, TXT, PDF, CSV file operations for the Agentic RAG system.
Supports read, write, extract, and convert operations.
"""

import os
import json
import logging
from typing import Dict, Any, Optional, List
from datetime import datetime

from .base_provider import BaseProvider, ProviderConfig, ProviderResult

logger = logging.getLogger(__name__)


class FileControlConfig(ProviderConfig):
    """Configuration for file control provider"""
    allowed_extensions: List[str] = [".txt", ".pdf", ".xlsx", ".xls", ".csv", ".json", ".md"]
    max_file_size_mb: int = 50
    temp_directory: str = "./temp"
    enable_ocr: bool = False


class FileControlProvider(BaseProvider):
    """
    MCP Provider for controlling Excel, TXT, PDF, and other file types.
    
    Capabilities:
    - Read/Write TXT, JSON, MD files
    - Read/Write Excel files (xlsx, xls, csv)
    - Extract text from PDF files
    - Convert between formats
    """
    
    def __init__(self, config: FileControlConfig = None):
        super().__init__(config or FileControlConfig())
        self.config: FileControlConfig = self.config
        
    async def initialize(self) -> bool:
        """Initialize the file control provider"""
        try:
            # Ensure temp directory exists
            os.makedirs(self.config.temp_directory, exist_ok=True)
            
            # Check required libraries
            self._check_dependencies()
            
            self._initialized = True
            logger.info("FileControlProvider initialized successfully")
            return True
            
        except Exception as e:
            logger.error(f"Failed to initialize FileControlProvider: {e}")
            return False
    
    def _check_dependencies(self):
        """Check if required libraries are available"""
        self._has_openpyxl = False
        self._has_pypdf = False
        self._has_pandas = False
        
        try:
            import openpyxl
            self._has_openpyxl = True
        except ImportError:
            logger.warning("openpyxl not installed - Excel support limited")
            
        try:
            from pypdf import PdfReader
            self._has_pypdf = True
        except ImportError:
            logger.warning("pypdf not installed - PDF support limited")
            
        try:
            import pandas as pd
            self._has_pandas = True
        except ImportError:
            logger.warning("pandas not installed - DataFrame support limited")
    
    async def health_check(self) -> bool:
        """Check if provider is healthy"""
        self._is_healthy = os.path.isdir(self.config.temp_directory)
        self._last_health_check = datetime.now()
        return self._is_healthy
    
    def get_capabilities(self) -> List[str]:
        """List available operations"""
        caps = ["read_txt", "write_txt", "read_json", "write_json", "read_md"]
        if self._has_openpyxl:
            caps.extend(["read_excel", "write_excel"])
        if self._has_pypdf:
            caps.extend(["read_pdf", "extract_pdf_text"])
        if self._has_pandas:
            caps.extend(["read_csv", "write_csv", "excel_to_csv", "csv_to_excel"])
        return caps
    
    # ==================== TXT Operations ====================
    
    async def read_txt(self, file_path: str, encoding: str = "utf-8") -> ProviderResult:
        """Read content from a text file"""
        try:
            if not os.path.exists(file_path):
                return ProviderResult(
                    success=False,
                    error=f"File not found: {file_path}",
                    provider=self.provider_name,
                    operation="read_txt"
                )
            
            with open(file_path, 'r', encoding=encoding) as f:
                content = f.read()
            
            return ProviderResult(
                success=True,
                data={"content": content, "size": len(content), "path": file_path},
                provider=self.provider_name,
                operation="read_txt"
            )
            
        except Exception as e:
            logger.error(f"Error reading TXT file: {e}")
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="read_txt"
            )
    
    async def write_txt(self, file_path: str, content: str, encoding: str = "utf-8", append: bool = False) -> ProviderResult:
        """Write content to a text file"""
        try:
            mode = 'a' if append else 'w'
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)
            
            with open(file_path, mode, encoding=encoding) as f:
                f.write(content)
            
            return ProviderResult(
                success=True,
                data={"path": file_path, "size": len(content), "mode": mode},
                provider=self.provider_name,
                operation="write_txt"
            )
            
        except Exception as e:
            logger.error(f"Error writing TXT file: {e}")
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="write_txt"
            )
    
    # ==================== JSON Operations ====================
    
    async def read_json(self, file_path: str) -> ProviderResult:
        """Read and parse a JSON file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            return ProviderResult(
                success=True,
                data=data,
                provider=self.provider_name,
                operation="read_json"
            )
            
        except json.JSONDecodeError as e:
            return ProviderResult(
                success=False,
                error=f"Invalid JSON: {e}",
                provider=self.provider_name,
                operation="read_json"
            )
        except Exception as e:
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="read_json"
            )
    
    async def write_json(self, file_path: str, data: Any, indent: int = 2) -> ProviderResult:
        """Write data to a JSON file"""
        try:
            os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=indent, ensure_ascii=False, default=str)
            
            return ProviderResult(
                success=True,
                data={"path": file_path},
                provider=self.provider_name,
                operation="write_json"
            )
            
        except Exception as e:
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="write_json"
            )
    
    # ==================== Excel Operations ====================
    
    async def read_excel(self, file_path: str, sheet_name: str = None) -> ProviderResult:
        """Read data from an Excel file"""
        if not self._has_openpyxl:
            return ProviderResult(
                success=False,
                error="openpyxl not installed. Run: pip install openpyxl",
                provider=self.provider_name,
                operation="read_excel"
            )
        
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return ProviderResult(
                        success=False,
                        error=f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}",
                        provider=self.provider_name,
                        operation="read_excel"
                    )
                sheet = wb[sheet_name]
            else:
                sheet = wb.active
            
            # Convert to list of dictionaries
            data = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_data[headers[i]] = value
                if any(v is not None for v in row_data.values()):
                    data.append(row_data)
            
            wb.close()
            
            return ProviderResult(
                success=True,
                data={
                    "rows": data,
                    "row_count": len(data),
                    "headers": headers,
                    "sheets": wb.sheetnames,
                    "path": file_path
                },
                provider=self.provider_name,
                operation="read_excel"
            )
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="read_excel"
            )
    
    async def write_excel(self, file_path: str, data: List[Dict], sheet_name: str = "Sheet1") -> ProviderResult:
        """Write data to an Excel file"""
        if not self._has_openpyxl:
            return ProviderResult(
                success=False,
                error="openpyxl not installed",
                provider=self.provider_name,
                operation="write_excel"
            )
        
        try:
            import openpyxl
            
            os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = sheet_name
            
            if data:
                # Write headers
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
                
                # Write data rows
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            wb.save(file_path)
            wb.close()
            
            return ProviderResult(
                success=True,
                data={"path": file_path, "rows_written": len(data)},
                provider=self.provider_name,
                operation="write_excel"
            )
            
        except Exception as e:
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="write_excel"
            )
    
    # ==================== PDF Operations ====================
    
    async def read_pdf(self, file_path: str, pages: List[int] = None) -> ProviderResult:
        """Extract text from a PDF file"""
        if not self._has_pypdf:
            return ProviderResult(
                success=False,
                error="pypdf not installed. Run: pip install pypdf",
                provider=self.provider_name,
                operation="read_pdf"
            )
        
        try:
            from pypdf import PdfReader
            
            reader = PdfReader(file_path)
            total_pages = len(reader.pages)
            
            text_by_page = {}
            pages_to_read = pages or range(total_pages)
            
            for page_num in pages_to_read:
                if 0 <= page_num < total_pages:
                    text = reader.pages[page_num].extract_text() or ""
                    text_by_page[page_num] = text
            
            full_text = "\n\n".join([text_by_page[p] for p in sorted(text_by_page.keys())])
            
            return ProviderResult(
                success=True,
                data={
                    "text": full_text,
                    "pages": text_by_page,
                    "total_pages": total_pages,
                    "path": file_path,
                    "metadata": {
                        "author": reader.metadata.author if reader.metadata else None,
                        "title": reader.metadata.title if reader.metadata else None
                    }
                },
                provider=self.provider_name,
                operation="read_pdf"
            )
            
        except Exception as e:
            logger.error(f"Error reading PDF: {e}")
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="read_pdf"
            )
    
    # ==================== CSV Operations ====================
    
    async def read_csv(self, file_path: str, delimiter: str = ",") -> ProviderResult:
        """Read data from a CSV file"""
        try:
            import csv
            
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                data = list(reader)
                headers = reader.fieldnames
            
            return ProviderResult(
                success=True,
                data={
                    "rows": data,
                    "row_count": len(data),
                    "headers": headers,
                    "path": file_path
                },
                provider=self.provider_name,
                operation="read_csv"
            )
            
        except Exception as e:
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="read_csv"
            )
    
    async def write_csv(self, file_path: str, data: List[Dict], delimiter: str = ",") -> ProviderResult:
        """Write data to a CSV file"""
        try:
            import csv
            
            os.makedirs(os.path.dirname(file_path) or '.', exist_ok=True)
            
            if not data:
                return ProviderResult(
                    success=False,
                    error="No data to write",
                    provider=self.provider_name,
                    operation="write_csv"
                )
            
            headers = list(data[0].keys())
            
            with open(file_path, 'w', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=headers, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(data)
            
            return ProviderResult(
                success=True,
                data={"path": file_path, "rows_written": len(data)},
                provider=self.provider_name,
                operation="write_csv"
            )
            
        except Exception as e:
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="write_csv"
            )
    
    # ==================== Utility Operations ====================
    
    async def list_files(self, directory: str, extensions: List[str] = None) -> ProviderResult:
        """List files in a directory"""
        try:
            if not os.path.isdir(directory):
                return ProviderResult(
                    success=False,
                    error=f"Directory not found: {directory}",
                    provider=self.provider_name,
                    operation="list_files"
                )
            
            files = []
            for item in os.listdir(directory):
                full_path = os.path.join(directory, item)
                if os.path.isfile(full_path):
                    ext = os.path.splitext(item)[1].lower()
                    if extensions is None or ext in extensions:
                        stat = os.stat(full_path)
                        files.append({
                            "name": item,
                            "path": full_path,
                            "extension": ext,
                            "size": stat.st_size,
                            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
                        })
            
            return ProviderResult(
                success=True,
                data={"files": files, "count": len(files), "directory": directory},
                provider=self.provider_name,
                operation="list_files"
            )
            
        except Exception as e:
            return ProviderResult(
                success=False,
                error=str(e),
                provider=self.provider_name,
                operation="list_files"
            )


# Singleton instance
file_control_provider = FileControlProvider()
